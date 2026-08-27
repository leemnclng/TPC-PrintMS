from __future__ import annotations

from io import BytesIO
from statistics import mean

from openpyxl import load_workbook

from ..models.document_analysis import DocumentAnalysis, PageMargins
from ..models.enums import DocumentFileType, Orientation, PaperSize
from ..utils.color_analysis import ImageColorMetrics, analyze_image_color, estimate_text_ink_coverage
from ..utils.image_processing import open_image
from .base import DocumentAnalyzer, InvalidDocumentError, estimate_print_time

PAPER_SIZE_BY_CODE = {
    "8": PaperSize.a3,
    "9": PaperSize.a4,
    "1": PaperSize.letter,
    "5": PaperSize.legal,
}
MAX_ROWS_ANALYZED = 5_000
MAX_COLUMNS_ANALYZED = 200


class ExcelAnalyzer(DocumentAnalyzer):
    file_type = DocumentFileType.xlsx

    def analyze(self, filename: str, data: bytes, mime_type: str) -> DocumentAnalysis:
        try:
            workbook = load_workbook(BytesIO(data), data_only=True, read_only=False)
        except Exception as error:
            raise InvalidDocumentError("The selected Excel workbook is damaged or unsupported.") from error

        sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        printable_sheets = [sheet for sheet in sheets if self._has_content(sheet)] or sheets[:1]
        if not printable_sheets:
            raise InvalidDocumentError("The selected workbook has no visible worksheets.")

        text_values: list[str] = []
        table_count = 0
        graphic_count = 0
        image_count = 0
        image_metrics: list[ImageColorMetrics] = []
        color_pages = 0
        orientations: list[Orientation] = []
        paper_sizes: list[PaperSize] = []
        truncated = False

        for sheet in printable_sheets:
            sheet_colored = False
            max_row = min(sheet.max_row, MAX_ROWS_ANALYZED)
            max_column = min(sheet.max_column, MAX_COLUMNS_ANALYZED)
            truncated = truncated or sheet.max_row > max_row or sheet.max_column > max_column
            for row in sheet.iter_rows(max_row=max_row, max_col=max_column):
                for cell in row:
                    if cell.value is not None:
                        text_values.append(str(cell.value))
                    if self._cell_is_colored(cell):
                        sheet_colored = True
            table_count += len(sheet.tables)
            graphic_count += len(sheet._charts)
            for image in sheet._images:
                image_count += 1
                try:
                    metric = analyze_image_color(open_image(image._data()))
                    image_metrics.append(metric)
                    sheet_colored = sheet_colored or metric.is_colored
                except Exception:
                    pass
            color_pages += int(sheet_colored)
            orientations.append(
                Orientation.landscape if sheet.page_setup.orientation == "landscape" else Orientation.portrait
            )
            paper_sizes.append(PAPER_SIZE_BY_CODE.get(str(sheet.page_setup.paperSize), PaperSize.unknown))

        text = " ".join(text_values)
        page_count = len(printable_sheets)
        bw_pages = page_count - color_pages
        first_sheet = printable_sheets[0]
        orientation = orientations[0] if len(set(orientations)) == 1 else Orientation.mixed
        paper_size = paper_sizes[0] if len(set(paper_sizes)) == 1 else PaperSize.unknown
        coverage = round(mean(metric.coverage_percent for metric in image_metrics), 1) if image_metrics else 0.0
        text_ink = estimate_text_ink_coverage(len(text), page_count)
        image_ink = mean(metric.ink_coverage_percent for metric in image_metrics) if image_metrics else 0.0
        image_color = mean(metric.color_coverage_percent for metric in image_metrics) if image_metrics else 0.0
        ink = round(min(100.0, text_ink + image_ink), 1)
        color_coverage = round(
            min(100.0, image_color + text_ink * (color_pages / max(page_count, 1))),
            1,
        )
        warnings = [
            "XLSX pagination is estimated as one page per non-empty visible worksheet; print areas and scaling can change the final count."
        ]
        if truncated:
            warnings.append("Cell inspection was capped for safety; counts may exclude content beyond the analyzed range.")
        warnings.append("Ink coverage is estimated from cell density and embedded images because XLSX is not rendered.")
        return DocumentAnalysis(
            filename=filename,
            file_type=self.file_type,
            mime_type=mime_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_size_bytes=len(data),
            page_count=page_count,
            paper_size=paper_size,
            orientation=orientation,
            width_mm=None,
            height_mm=None,
            dpi=None,
            character_count=len(text),
            word_count=len(text.split()),
            ocr_required=False,
            image_count=image_count,
            contains_images=image_count > 0,
            image_coverage_percent=coverage,
            estimated_color_coverage_percent=color_coverage,
            estimated_ink_coverage_percent=ink,
            table_count=table_count,
            graphic_count=graphic_count,
            margins=PageMargins(
                top_mm=round(first_sheet.page_margins.top * 25.4, 1),
                right_mm=round(first_sheet.page_margins.right * 25.4, 1),
                bottom_mm=round(first_sheet.page_margins.bottom * 25.4, 1),
                left_mm=round(first_sheet.page_margins.left * 25.4, 1),
            ),
            color_pages=color_pages,
            bw_pages=bw_pages,
            duplex_compatible=page_count > 1,
            estimated_print_time_seconds=estimate_print_time(color_pages, bw_pages),
            confidence=0.68,
            warnings=warnings,
        )

    @staticmethod
    def _has_content(sheet) -> bool:
        max_row = min(sheet.max_row, MAX_ROWS_ANALYZED)
        max_column = min(sheet.max_column, MAX_COLUMNS_ANALYZED)
        return any(
            cell.value is not None
            for row in sheet.iter_rows(max_row=max_row, max_col=max_column)
            for cell in row
        )

    @staticmethod
    def _cell_is_colored(cell) -> bool:
        for color in (cell.font.color, cell.fill.fgColor):
            if not color or color.type != "rgb" or not color.rgb:
                continue
            raw = str(color.rgb)[-6:]
            try:
                values = [int(raw[index:index + 2], 16) for index in (0, 2, 4)]
            except ValueError:
                continue
            if max(values) - min(values) > 14:
                return True
        return False
